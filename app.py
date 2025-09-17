from flask import Flask, render_template, request, redirect, Response
import pandas as pd
import os
import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill

app = Flask(__name__)

ARQUIVO = "solicitacoes_manutencao.xlsx"

# Configurações do e-mail (Locaweb)
SMTP_SERVER = "smtp.majonav.com.br"
SMTP_PORT = 587
EMAIL_REMETENTE = "jonas.santos@majonav.com.br"
EMAIL_SENHA = "m@j0n@v&JS"
EMAIL_TECNICO = "jonas.santos@majonav.com.br"

# ---------------- Funções auxiliares ---------------- #

def carregar_planilha():
    if os.path.exists(ARQUIVO):
        return pd.read_excel(ARQUIVO)
    else:
        return pd.DataFrame(columns=["ID", "Solicitante", "E-mail", "Equipamento", "Descrição", "Data", "Status"])

def salvar_planilha(df):
    df.to_excel(ARQUIVO, index=False)

    wb = load_workbook(ARQUIVO)
    ws = wb.active

    ultima_linha = ws.max_row
    ultima_coluna = ws.max_column
    col_final = chr(64 + ultima_coluna)

    range_tabela = f"A1:{col_final}{ultima_linha}"

    if ws._tables:
        ws._tables.clear()

    tabela = Table(displayName="Solicitacoes", ref=range_tabela)
    estilo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    # Cores na coluna Status
    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    status_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Status":
            status_col = idx
            break

    if status_col:
        for row in range(2, ultima_linha + 1):
            status_cell = ws.cell(row=row, column=status_col)
            if status_cell.value == "Concluída":
                status_cell.fill = verde
            elif status_cell.value == "Aberta":
                status_cell.fill = vermelho

    wb.save(ARQUIVO)

def enviar_email(solicitante, email_solicitante, equipamento, descricao, data):
    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_REMETENTE
        msg["To"] = EMAIL_TECNICO
        msg["Subject"] = "📢 Nova Solicitação de Manutenção Preventiva"
        msg["Reply-To"] = email_solicitante

        corpo = f"""
Olá, técnico responsável!

Uma nova solicitação de manutenção preventiva foi registrada:

🧑 Solicitante: {solicitante} ({email_solicitante})
🛠️ Equipamento: {equipamento}
📝 Descrição: {descricao}
📅 Data: {data}

Por favor, verifique o sistema e agende a manutenção.
"""
        msg.attach(MIMEText(corpo, "plain"))

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_REMETENTE, EMAIL_SENHA)
        server.sendmail(EMAIL_REMETENTE, EMAIL_TECNICO, msg.as_string())
        server.quit()

        print("✅ E-mail enviado com sucesso!")

    except Exception as e:
        print("❌ Erro ao enviar e-mail:", e)

# ---------------- Rotas ---------------- #

@app.route("/")
def index():
    return render_template("formulario.html")

@app.route("/nova", methods=["POST"])
def nova():
    df = carregar_planilha()
    novo_id = len(df) + 1
    solicitante = request.form["solicitante"]
    email_solicitante = request.form["email_solicitante"]
    equipamento = request.form["equipamento"]
    descricao = request.form["descricao"]
    data = datetime.datetime.now().strftime("%d/%m/%Y")  # somente data, sem hora

    nova = {
        "ID": novo_id,
        "Solicitante": solicitante,
        "E-mail": email_solicitante,
        "Equipamento": equipamento,
        "Descrição": descricao,
        "Data": data,
        "Status": "Aberta",
    }

    df = pd.concat([df, pd.DataFrame([nova])], ignore_index=True)
    salvar_planilha(df)

    enviar_email(solicitante, email_solicitante, equipamento, descricao, data)

    return redirect("/")

# ---------------- Login técnico ---------------- #

USUARIO = "tecnico"
SENHA = "1234"

def autenticar(usuario, senha):
    return usuario == USUARIO and senha == SENHA

def exigir_autenticacao():
    return Response(
        "Acesso restrito. Faça login como técnico.",
        401,
        {"WWW-Authenticate": 'Basic realm="Login técnico"'},
    )

@app.route("/registros")
def registros():
    auth = request.authorization
    if not auth or not autenticar(auth.username, auth.password):
        return exigir_autenticacao()

    df = carregar_planilha()
    return render_template("registros.html", solicitacoes=df.to_dict(orient="records"))

@app.route("/concluir/<int:id_solicitacao>", methods=["POST"])
def concluir(id_solicitacao):
    auth = request.authorization
    if not auth or not autenticar(auth.username, auth.password):
        return exigir_autenticacao()

    df = carregar_planilha()
    df.loc[df["ID"] == id_solicitacao, "Status"] = "Concluída"
    salvar_planilha(df)

    return redirect("/registros")

# ---------------- Rodar aplicação ---------------- #

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080, debug=True)
